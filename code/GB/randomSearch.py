import os
import pandas as pd
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import warnings

# 경고 메시지 무시 설정
warnings.filterwarnings('ignore')

# 데이터 폴더 경로 지정
folder_path = 'C:/Users/yujin/Desktop/work/data/Preprocessed/Step2_Balanced/'
output_file = 'C:/Users/yujin/Desktop/work/result/gradient_boosting_random_search_results.xlsx'

# 폴더 내의 모든 CSV 파일에 대해 반복 수행
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        dataset_name = filename.split('_')[0]  # 파일명에서 데이터셋 이름 추출 (예: CM1, JM1 등)
        file_path = os.path.join(folder_path, filename)

        # 데이터 로드
        data = pd.read_csv(file_path)

        # 특성과 라벨 분리
        X = data.drop('Defective', axis=1)
        y = data['Defective']

        # 데이터셋 분할 (train/test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # 그라디언트 부스트 모델 초기화
        gb = GradientBoostingClassifier(random_state=42)

        # 하이퍼파라미터 랜덤 서치 설정
        param_distributions = {
            'n_estimators': [50, 100, 200, 300],
            'max_depth': [3, 5, 10, 15],
            'min_samples_split': [2, 5, 10],
            'min_samples_leaf': [1, 2, 4],
            'learning_rate': [0.01, 0.05, 0.1, 0.2]
        }

        # 랜덤 서치 수행
        random_search = RandomizedSearchCV(estimator=gb, param_distributions=param_distributions, n_iter=30, cv=5, n_jobs=-1, verbose=2, scoring='f1', random_state=42)
        random_search.fit(X_train, y_train)

        # 최적의 하이퍼파라미터 추출
        best_params = random_search.best_params_
        n_estimators = best_params['n_estimators']
        max_depth = best_params['max_depth']
        min_samples_split = best_params['min_samples_split']
        min_samples_leaf = best_params['min_samples_leaf']
        learning_rate = best_params['learning_rate']

        # 최적의 모델로 예측 수행
        best_gb = random_search.best_estimator_
        y_pred = best_gb.predict(X_test)
        y_pred_proba = best_gb.predict_proba(X_test)[:, 1]

        # 성능 지표 계산
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average='binary')
        recall = recall_score(y_test, y_pred, average='binary')
        f1 = f1_score(y_test, y_pred, average='binary')
        auc = roc_auc_score(y_test, y_pred_proba)

        # 분류 리포트 생성
        report = classification_report(y_test, y_pred, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        print(f"\n[{dataset_name}] 성능 지표:")
        print(f"정확도: {accuracy:.4f}")
        print(f"정밀도: {precision:.4f}")
        print(f"재현율: {recall:.4f}")
        print(f"F1 스코어: {f1:.4f}")
        print(f"AUC: {auc:.4f}")

        # 최적 하이퍼파라미터 저장
        params_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Parameter': ['n_estimators', 'max_depth', 'min_samples_split', 'min_samples_leaf', 'learning_rate'],
            'Best Value': [n_estimators, max_depth, min_samples_split, min_samples_leaf, learning_rate]
        })

        # 엑셀 파일에 결과 저장
        # 데이터프레임 생성 (AUC 값 추가)
        results_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Value': [accuracy, precision, recall, f1, auc]
        })

        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            # 파일이 존재하지 않으면 새로운 파일 생성 및 데이터 저장
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)
        else:
            # 파일이 존재하면 새로운 시트로 추가 저장
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                if f'{dataset_name} Performance Metrics' in writer.book.sheetnames:
                    writer.book.remove(writer.book[f'{dataset_name} Performance Metrics'])
                if f'{dataset_name} Best Hyperparameters' in writer.book.sheetnames:
                    writer.book.remove(writer.book[f'{dataset_name} Best Hyperparameters'])
                if f'{dataset_name} Classification Report' in writer.book.sheetnames:
                    writer.book.remove(writer.book[f'{dataset_name} Classification Report'])
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)

        print(f"엑셀 파일이 성공적으로 업데이트되었습니다: {dataset_name}")
