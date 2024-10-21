import os
import pandas as pd
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 데이터 폴더 경로 지정
folder_path = 'C:/Users/yujin/Desktop/work/data/Preprocessed/Step2_Balanced/'
output_file = 'C:/Users/yujin/Desktop/work/result/extra_trees_grid_search_results.xlsx'

# 폴더 내의 모든 CSV 파일에 대해 반복 수행
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        dataset_name = filename.split('_')[0]  # 파일명에서 데이터셋 이름 추출 (예: CM1, JM1 등)
        file_path = os.path.join(folder_path, filename)

        # 데이터 로드
        data = pd.read_csv(file_path)

        # 특성과 라벨 분리
        X = data.drop('Defective', axis=1)
        y = data['Defective']

        # 데이터셋 분할 (train/test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # 엑스트라 트리 모델 초기화
        etc = ExtraTreesClassifier(random_state=42)

        # 하이퍼파라미터 그리드 설정
        param_grid = {
            'n_estimators': [50, 100, 200],
            'max_depth': [None, 10, 20, 30],
            'min_samples_split': [2, 5, 10],
            'min_samples_leaf': [1, 2, 4],
            'bootstrap': [True, False]
        }

        # 그리드 서치 수행
        grid_search = GridSearchCV(estimator=etc, param_grid=param_grid, cv=5, n_jobs=-1, verbose=2, scoring='f1')
        grid_search.fit(X_train, y_train)

        # 최적의 하이퍼파라미터 출력
        best_params = grid_search.best_params_
        print(f"\n최적의 하이퍼파라미터 조합 ({dataset_name}):")
        print(best_params)

        # 최적의 모델로 예측 수행
        best_etc = grid_search.best_estimator_
        y_pred = best_etc.predict(X_test)
        y_pred_proba = best_etc.predict_proba(X_test)[:, 1]

        # 성능 지표 계산
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average='binary')
        recall = recall_score(y_test, y_pred, average='binary')
        f1 = f1_score(y_test, y_pred, average='binary')
        auc = roc_auc_score(y_test, y_pred_proba)

        # 분류 리포트 생성
        report = classification_report(y_test, y_pred, output_dict=True)
        report_df = pd.DataFrame(report).transpose()

        print("\n성능 지표:")
        print(f"정확도: {accuracy:.4f}")
        print(f"정밀도: {precision:.4f}")
        print(f"재현율: {recall:.4f}")
        print(f"F1 스코어: {f1:.4f}")
        print(f"AUC: {auc:.4f}")

        # 데이터프레임 생성 (AUC 값 추가)
        results_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Value': [accuracy, precision, recall, f1, auc]
        })

        params_df = pd.DataFrame({
            'Dataset': [dataset_name] * len(best_params),
            'Parameter': list(best_params.keys()),
            'Best Value': list(best_params.values())
        })

        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            # 파일이 존재하지 않으면 새로운 파일 생성 및 데이터 저장
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name=f'{dataset_name}_Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name}_Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name}_Classification Report', index=True)
        else:
            # 파일이 존재하면 openpyxl로 불러오기
            book = load_workbook(output_file)

            # 'Performance Metrics' 시트에 데이터 추가 또는 생성
            performance_metrics_sheet = f'{dataset_name}_Performance Metrics'
            if performance_metrics_sheet in book.sheetnames:
                sheet = book[performance_metrics_sheet]
                for r in dataframe_to_rows(results_df, index=False, header=False):
                    sheet.append(r)
            else:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                    results_df.to_excel(writer, sheet_name=performance_metrics_sheet, index=False)

            # 'Best Hyperparameters' 시트에 데이터 추가 또는 생성
            best_hyperparams_sheet = f'{dataset_name}_Best Hyperparameters'
            if best_hyperparams_sheet in book.sheetnames:
                sheet = book[best_hyperparams_sheet]
                for r in dataframe_to_rows(params_df, index=False, header=False):
                    sheet.append(r)
            else:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                    params_df.to_excel(writer, sheet_name=best_hyperparams_sheet, index=False)

            # 'Classification Report' 시트에 데이터 추가 또는 생성
            classification_report_sheet = f'{dataset_name}_Classification Report'
            if classification_report_sheet in book.sheetnames:
                sheet = book[classification_report_sheet]
                for r in dataframe_to_rows(report_df, index=False, header=False):
                    sheet.append(r)
            else:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                    report_df.to_excel(writer, sheet_name=classification_report_sheet, index=True)

            # 저장
            book.save(output_file)

        print(f"엑셀 파일이 {dataset_name} 결과로 성공적으로 업데이트되었습니다.")
