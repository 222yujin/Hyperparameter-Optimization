import pandas as pd
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 데이터 로드
dataset_name = 'PC4'  # 데이터셋 이름 지정 CM1 JM1 KC1 KC3 MC1 MC2 MW1 PC1 PC2 PC3 PC4
file_path = f'C:/Users/yujin/Desktop/work/data/Preprocessed/Step2_Balanced/{dataset_name}_Clean_Balanced.csv'
data = pd.read_csv(file_path)


# 특성과 라벨 분리
X = data.drop('Defective', axis=1)
y = data['Defective']

# 데이터셋 분할 (train/test)
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# 엑스트라 트리 모델 초기화
et = ExtraTreesClassifier(random_state=42)

# 하이퍼파라미터 랜덤 서치 설정
param_distributions = {
    'n_estimators': [50, 100, 200, 300],
    'max_depth': [None, 10, 20, 30, 40],
    'min_samples_split': [2, 5, 10],
    'min_samples_leaf': [1, 2, 4],
    'bootstrap': [True, False]
}

# 랜덤 서치 수행
random_search = RandomizedSearchCV(estimator=et, param_distributions=param_distributions, n_iter=30, cv=5, n_jobs=-1, verbose=2, scoring='f1', random_state=42)
random_search.fit(X_train, y_train)

# 최적의 하이퍼파라미터 출력
best_params = random_search.best_params_
print("\n최적의 하이퍼파라미터 조합:")
print(best_params)

# 최적의 모델로 예측 수행
best_et = random_search.best_estimator_
y_pred = best_et.predict(X_test)
y_pred_proba = best_et.predict_proba(X_test)[:, 1]

# 성능 지표 계산
accuracy = accuracy_score(y_test, y_pred)
precision = precision_score(y_test, y_pred, average='binary')
recall = recall_score(y_test, y_pred, average='binary')
f1 = f1_score(y_test, y_pred, average='binary')
auc = roc_auc_score(y_test, y_pred_proba)

# 분류 리포트 생성
report = classification_report(y_test, y_pred, output_dict=True)
report_df = pd.DataFrame(report).transpose()

print("\n성능 지표:")
print(f"정확도: {accuracy:.4f}")
print(f"정밀도: {precision:.4f}")
print(f"재현율: {recall:.4f}")
print(f"F1 스코어: {f1:.4f}")
print(f"AUC: {auc:.4f}")

# 엑셀 파일에 결과 저장
output_file = 'C:/Users/yujin/Desktop/work/result/extra_trees_random_search_results.xlsx'

# 데이터프레임 생성 (AUC 값 추가)
results_df = pd.DataFrame({
    'Dataset': [dataset_name] * 5,
    'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
    'Value': [accuracy, precision, recall, f1, auc]
})

params_df = pd.DataFrame({
    'Dataset': [dataset_name] * len(best_params),
    'Parameter': list(best_params.keys()),
    'Best Value': list(best_params.values())
})

# 엑셀 파일이 존재하는지 확인
if not os.path.exists(output_file):
    # 파일이 존재하지 않으면 새로운 파일 생성 및 데이터 저장
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Performance Metrics', index=False)
        params_df.to_excel(writer, sheet_name='Best Hyperparameters', index=False)
        report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)
else:
    # 파일이 존재하면 openpyxl로 불러오기
    book = load_workbook(output_file)

    # 'Performance Metrics' 시트에 데이터 추가
    if 'Performance Metrics' in book.sheetnames:
        sheet = book['Performance Metrics']
        for r in dataframe_to_rows(results_df, index=False, header=False):
            sheet.append(r)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            results_df.to_excel(writer, sheet_name='Performance Metrics', index=False)

    # 'Best Hyperparameters' 시트에 데이터 추가
    if 'Best Hyperparameters' in book.sheetnames:
        sheet = book['Best Hyperparameters']
        for r in dataframe_to_rows(params_df, index=False, header=False):
            sheet.append(r)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            params_df.to_excel(writer, sheet_name='Best Hyperparameters', index=False)

    # 'Classification Report' 시트에 데이터 추가
    if f'{dataset_name} Classification Report' in book.sheetnames:
        sheet = book[f'{dataset_name} Classification Report']
        for r in dataframe_to_rows(report_df, index=False, header=False):
            sheet.append(r)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)

    # 저장
    book.save(output_file)

print("엑셀 파일이 성공적으로 업데이트되었습니다.")
