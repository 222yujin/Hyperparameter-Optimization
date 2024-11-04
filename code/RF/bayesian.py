import os
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from bayes_opt import BayesianOptimization
from openpyxl import load_workbook
import warnings

# 경고 메시지 무시 설정
warnings.filterwarnings('ignore')

# 데이터 폴더 경로 지정
folder_path = 'C:/Users/user/Desktop/work/data/Preprocessed/Step2_Balanced/'
output_file = 'C:/Users/user/Desktop/work/result/random_forest_bayesian_optimization_results.xlsx'

# 폴더 내의 모든 CSV 파일에 대해 반복 수행
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        dataset_name = filename.split('_')[0]  # 파일명에서 데이터셋 이름 추출
        file_path = os.path.join(folder_path, filename)

        # 데이터 로드
        data = pd.read_csv(file_path)

        # 특성과 라벨 분리
        X = data.drop('Defective', axis=1)
        y = data['Defective']

        # 데이터셋 분할 (train/test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # 베이지안 최적화를 위한 목적 함수 정의
        def rf_evaluate(n_estimators, max_depth, min_samples_split, min_samples_leaf, max_features):
            model = RandomForestClassifier(
                n_estimators=int(n_estimators),
                max_depth=int(max_depth),
                min_samples_split=int(min_samples_split),
                min_samples_leaf=int(min_samples_leaf),
                max_features=max(min(max_features, 0.999), 0.001),  # max_features가 1을 넘지 않도록 제한
                random_state=42
            )
            cv_score = cross_val_score(model, X_train, y_train, cv=5, scoring='f1').mean()
            return cv_score

        # 베이지안 최적화 설정
        pbounds = {
            'n_estimators': (50, 300),
            'max_depth': (5, 15),
            'min_samples_split': (2, 10),
            'min_samples_leaf': (1, 4),
            'max_features': (0.1, 0.9)
        }

        optimizer = BayesianOptimization(f=rf_evaluate, pbounds=pbounds, random_state=42)
        optimizer.maximize(init_points=5, n_iter=20)

        # 최적의 하이퍼파라미터 추출
        best_params = optimizer.max['params']
        n_estimators = int(best_params['n_estimators'])
        max_depth = int(best_params['max_depth'])
        min_samples_split = int(best_params['min_samples_split'])
        min_samples_leaf = int(best_params['min_samples_leaf'])
        max_features = best_params['max_features']

        # 최적의 모델로 예측 수행
        best_rf = RandomForestClassifier(
            n_estimators=n_estimators,
            max_depth=max_depth,
            min_samples_split=min_samples_split,
            min_samples_leaf=min_samples_leaf,
            max_features=max_features,
            random_state=42
        )
        best_rf.fit(X_train, y_train)
        y_pred = best_rf.predict(X_test)
        y_pred_proba = best_rf.predict_proba(X_test)[:, 1]

        # 성능 지표 계산
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average='binary')
        recall = recall_score(y_test, y_pred, average='binary')
        f1 = f1_score(y_test, y_pred, average='binary')
        auc = roc_auc_score(y_test, y_pred_proba)

        # 분류 리포트 생성
        report = classification_report(y_test, y_pred, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        print(f"\n[{dataset_name}] 성능 지표:")
        print(f"정확도: {accuracy:.4f}")
        print(f"정밀도: {precision:.4f}")
        print(f"재현율: {recall:.4f}")
        print(f"F1 스코어: {f1:.4f}")
        print(f"AUC: {auc:.4f}")

        # 최적 하이퍼파라미터 저장
        params_df = pd.DataFrame({
            'Dataset': [dataset_name] * len(best_params),
            'Parameter': list(best_params.keys()),
            'Best Value': list(best_params.values())
        })

        # 엑셀 파일에 결과 저장
        results_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Value': [accuracy, precision, recall, f1, auc]
        })

        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)
        else:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                book = writer.book
                if f'{dataset_name} Performance Metrics' in book.sheetnames:
                    del book[f'{dataset_name} Performance Metrics']
                if f'{dataset_name} Best Hyperparameters' in book.sheetnames:
                    del book[f'{dataset_name} Best Hyperparameters']
                if f'{dataset_name} Classification Report' in book.sheetnames:
                    del book[f'{dataset_name} Classification Report']
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)

        print(f"엑셀 파일이 성공적으로 업데이트되었습니다: {dataset_name}")
