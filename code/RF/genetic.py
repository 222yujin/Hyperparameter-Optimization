import os
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from deap import base, creator, tools, algorithms
from openpyxl import load_workbook
import warnings
import random

# 경고 메시지 무시 설정
warnings.filterwarnings('ignore')

# 데이터 폴더 경로 지정
folder_path = 'C:/Users/user/Desktop/work/data/Preprocessed/Step2_Balanced/'
output_file = 'C:/Users/user/Desktop/work/result/random_forest_genetic_algorithm_results.xlsx'

# 유전 알고리즘 설정
POP_SIZE = 10         # 인구 수
N_GEN = 10            # 세대 수
CX_PB = 0.5           # 교차 확률
MUT_PB = 0.2          # 돌연변이 확률

# 폴더 내의 모든 CSV 파일에 대해 반복 수행
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        dataset_name = filename.split('_')[0]  # 파일명에서 데이터셋 이름 추출
        file_path = os.path.join(folder_path, filename)

        # 데이터 로드
        data = pd.read_csv(file_path)

        # 특성과 라벨 분리
        X = data.drop('Defective', axis=1)
        y = data['Defective']

        # 데이터셋 분할 (train/test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # 평가 함수 정의
        def evaluate(individual):
            n_estimators = int(individual[0])
            max_depth = int(individual[1])
            min_samples_split = max(2, int(individual[2]))  # 최소값을 2로 설정
            min_samples_leaf = max(1, int(individual[3]))   # 최소값을 1로 설정
            max_features = max(0.1, min(individual[4], 1.0))  # 0.1과 1.0 사이로 제한

            model = RandomForestClassifier(
                n_estimators=n_estimators,
                max_depth=max_depth,
                min_samples_split=min_samples_split,
                min_samples_leaf=min_samples_leaf,
                max_features=max_features,
                random_state=42
            )
            cv_score = cross_val_score(model, X_train, y_train, cv=5, scoring='f1').mean()
            return cv_score,

        # 유전자 알고리즘 설정
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", list, fitness=creator.FitnessMax)

        toolbox = base.Toolbox()
        toolbox.register("attr_int_n_estimators", random.randint, 50, 300)
        toolbox.register("attr_int_max_depth", random.randint, 5, 15)
        toolbox.register("attr_int_min_samples_split", random.randint, 2, 10)  # 최소값을 2로 설정
        toolbox.register("attr_int_min_samples_leaf", random.randint, 1, 4)    # 최소값을 1로 설정
        toolbox.register("attr_float_max_features", random.uniform, 0.1, 1.0)  # 0.1과 1.0 사이로 제한

        toolbox.register("individual", tools.initCycle, creator.Individual,
                         (toolbox.attr_int_n_estimators, toolbox.attr_int_max_depth,
                          toolbox.attr_int_min_samples_split, toolbox.attr_int_min_samples_leaf,
                          toolbox.attr_float_max_features), n=1)
        toolbox.register("population", tools.initRepeat, list, toolbox.individual)

        toolbox.register("evaluate", evaluate)
        toolbox.register("mate", tools.cxUniform, indpb=0.5)
        toolbox.register("mutate", tools.mutGaussian, mu=0, sigma=1, indpb=0.2)
        toolbox.register("select", tools.selTournament, tournsize=3)

        population = toolbox.population(n=POP_SIZE)
        hof = tools.HallOfFame(1)

        algorithms.eaSimple(population, toolbox, cxpb=CX_PB, mutpb=MUT_PB, ngen=N_GEN, halloffame=hof, verbose=True)

        # 최적의 하이퍼파라미터 추출
        best_individual = hof[0]
        n_estimators = int(best_individual[0])
        max_depth = int(best_individual[1])
        min_samples_split = max(2, int(best_individual[2]))  # 2 이상으로 제한
        min_samples_leaf = max(1, int(best_individual[3]))   # 1 이상으로 제한
        max_features = max(0.1, min(best_individual[4], 1.0))  # 0.1과 1.0 사이로 제한

        # 최적의 모델로 예측 수행
        best_rf = RandomForestClassifier(
            n_estimators=n_estimators,
            max_depth=max_depth,
            min_samples_split=min_samples_split,
            min_samples_leaf=min_samples_leaf,
            max_features=max_features,
            random_state=42
        )
        best_rf.fit(X_train, y_train)
        y_pred = best_rf.predict(X_test)
        y_pred_proba = best_rf.predict_proba(X_test)[:, 1]

        # 성능 지표 계산
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average='binary')
        recall = recall_score(y_test, y_pred, average='binary')
        f1 = f1_score(y_test, y_pred, average='binary')
        auc = roc_auc_score(y_test, y_pred_proba)

        # 분류 리포트 생성
        report = classification_report(y_test, y_pred, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        print(f"\n[{dataset_name}] 성능 지표:")
        print(f"정확도: {accuracy:.4f}")
        print(f"정밀도: {precision:.4f}")
        print(f"재현율: {recall:.4f}")
        print(f"F1 스코어: {f1:.4f}")
        print(f"AUC: {auc:.4f}")

        # 최적 하이퍼파라미터 저장
        params_df = pd.DataFrame({
            'Dataset': [dataset_name] * len(best_individual),
            'Parameter': ['n_estimators', 'max_depth', 'min_samples_split', 'min_samples_leaf', 'max_features'],
            'Best Value': best_individual
        })

        # 엑셀 파일에 결과 저장
        results_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Value': [accuracy, precision, recall, f1, auc]
        })

        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)
        else:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                book = writer.book
                if f'{dataset_name} Performance Metrics' in book.sheetnames:
                    del book[f'{dataset_name} Performance Metrics']
                if f'{dataset_name} Best Hyperparameters' in book.sheetnames:
                    del book[f'{dataset_name} Best Hyperparameters']
                if f'{dataset_name} Classification Report' in book.sheetnames:
                    del book[f'{dataset_name} Classification Report']
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)

        print(f"엑셀 파일이 성공적으로 업데이트되었습니다: {dataset_name}")
