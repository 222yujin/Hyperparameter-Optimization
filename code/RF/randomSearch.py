import os
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report, roc_auc_score
from openpyxl import load_workbook
import warnings

# 경고 메시지 무시 설정
warnings.filterwarnings('ignore')

# 데이터 폴더 경로 지정
folder_path = 'C:/Users/user/Desktop/work/data/Preprocessed/Step2_Balanced/'
output_file = 'C:/Users/user/Desktop/work/result/random_forest_random_search_results.xlsx'

# 폴더 내의 모든 CSV 파일에 대해 반복 수행
for filename in os.listdir(folder_path):
    if filename.endswith('.csv'):
        dataset_name = filename.split('_')[0]  # 파일명에서 데이터셋 이름 추출
        file_path = os.path.join(folder_path, filename)

        # 데이터 로드
        data = pd.read_csv(file_path)

        # 특성과 라벨 분리
        X = data.drop('Defective', axis=1)
        y = data['Defective']

        # 데이터셋 분할 (train/test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # 하이퍼파라미터 랜덤 서치 범위 설정
        param_distributions = {
            'n_estimators': np.arange(50, 301, 50),
            'max_depth': [None] + list(np.arange(5, 16, 5)),
            'min_samples_split': np.arange(2, 11, 3),
            'min_samples_leaf': np.arange(1, 5, 1),
            'max_features': ['auto', 'sqrt', 'log2']
        }

        # RandomizedSearchCV 설정
        random_search = RandomizedSearchCV(estimator=RandomForestClassifier(random_state=42),
                                           param_distributions=param_distributions,
                                           scoring='f1',  # F1 스코어를 최적화
                                           cv=5,
                                           n_iter=20,  # 랜덤하게 20개의 조합을 탐색
                                           n_jobs=-1,
                                           verbose=1,
                                           random_state=42)

        # 랜덤 서치 실행
        random_search.fit(X_train, y_train)

        # 최적의 하이퍼파라미터 추출
        best_params = random_search.best_params_

        # 최적의 모델로 예측 수행
        best_rf = random_search.best_estimator_
        y_pred = best_rf.predict(X_test)
        y_pred_proba = best_rf.predict_proba(X_test)[:, 1]

        # 성능 지표 계산
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average='binary')
        recall = recall_score(y_test, y_pred, average='binary')
        f1 = f1_score(y_test, y_pred, average='binary')
        auc = roc_auc_score(y_test, y_pred_proba)

        # 분류 리포트 생성
        report = classification_report(y_test, y_pred, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        print(f"\n[{dataset_name}] 성능 지표:")
        print(f"정확도: {accuracy:.4f}")
        print(f"정밀도: {precision:.4f}")
        print(f"재현율: {recall:.4f}")
        print(f"F1 스코어: {f1:.4f}")
        print(f"AUC: {auc:.4f}")

        # 최적 하이퍼파라미터 저장
        params_df = pd.DataFrame({
            'Dataset': [dataset_name] * len(best_params),
            'Parameter': list(best_params.keys()),
            'Best Value': list(best_params.values())
        })

        # 엑셀 파일에 결과 저장
        results_df = pd.DataFrame({
            'Dataset': [dataset_name] * 5,
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Value': [accuracy, precision, recall, f1, auc]
        })

        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)
        else:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                book = writer.book
                if f'{dataset_name} Performance Metrics' in book.sheetnames:
                    del book[f'{dataset_name} Performance Metrics']
                if f'{dataset_name} Best Hyperparameters' in book.sheetnames:
                    del book[f'{dataset_name} Best Hyperparameters']
                if f'{dataset_name} Classification Report' in book.sheetnames:
                    del book[f'{dataset_name} Classification Report']
                results_df.to_excel(writer, sheet_name=f'{dataset_name} Performance Metrics', index=False)
                params_df.to_excel(writer, sheet_name=f'{dataset_name} Best Hyperparameters', index=False)
                report_df.to_excel(writer, sheet_name=f'{dataset_name} Classification Report', index=True)

        print(f"엑셀 파일이 성공적으로 업데이트되었습니다: {dataset_name}")
